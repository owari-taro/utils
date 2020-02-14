import openpyxl as px
import os
from typing import List


def write_excel(year: int, month: int, csv_dir: str):
    """csv_fileをexcelファイルにまとめる"""
    csv_fnames = os.listdir(csv_dir)  # path
    wb = px.Workbook()
    for fname in csv_fnames:
        titles = wb.get_sheet_names()
        sheet = wb.get_sheet_by_name(titles[-1])
        sheet.title = fname.split(".")[0]
        csv_list: List[str] = load_csv(
            os.path.join(OUTPUT_DIR, fname))  # o#fname
        row_: int = 0
        col_: int = 0
        for row in csv_list:
            row_ += 1
            for ele in row:
                col_ += 1
                # print(ele)
                sheet.cell(row=row_, column=col_, value=ele)
            col_ = 0
        wb.create_sheet()

    excel_fname = "勤務表{}年{}月.xlsx".format(year, month)
    wb.save("勤務表{}年{}月.xlsx".format(year, month))
    print("勤務表を出力しました:{}\n".format(excel_fname))


if __name__ == "__main__":
    tmp: List[str] = []
    tmp.append("test")
    print(tmp)
